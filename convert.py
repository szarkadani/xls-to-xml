import openpyxl
from xml.etree.ElementTree import Element, SubElement, tostring

wb = openpyxl.load_workbook('file.xlsx')

sheet = wb.active

header = [cell.value for cell in sheet[1]]

root = Element('root')

parent_attr_count = 2
parent_tag_name = 'parent'
child_tag_name = 'child'

current_parent = root

for row in range(2, sheet.max_row + 1):
    parent_attrs_empty = all(
        cell.value is None for cell in sheet[row][:parent_attr_count])

    if not parent_attrs_empty:
        current_parent = SubElement(root, parent_tag_name)
        for col, attr in enumerate(header[:parent_attr_count]):
            cell_value = sheet.cell(row=row, column=col + 1).value
            if cell_value is not None:
                parent_attr = SubElement(current_parent, attr)
                parent_attr.text = str(cell_value)
    
    child = SubElement(current_parent, child_tag_name)
    for col, attr in enumerate(header[parent_attr_count:]):
        cell_value = sheet.cell(row=row, column=col +
                                parent_attr_count + 1).value
        if cell_value is not None:
            child_attr = SubElement(child, attr)
            child_attr.text = str(cell_value)

xml_str = tostring(root, encoding='unicode')
with open('xml_file.xml', 'w') as f:
    f.write(xml_str)
